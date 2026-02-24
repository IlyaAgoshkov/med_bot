from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, FSInputFile
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.exceptions import TelegramBadRequest
from app.database.db import get_db_path
from app.filters.admin_filter import IsAdminFilter
from app.keyboards.inline_keyboards import get_admin_keyboard, get_gift_keyboard
from app.states.survey_states import SurveyStates
import aiosqlite
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import asyncio
import tempfile
import os
import logging
from datetime import datetime

logger = logging.getLogger(__name__)
router = Router()


def _excel_safe_value(value):
    """Экранировать строки, начинающиеся с =, +, -, @, чтобы Excel не воспринимал их как формулу."""
    if value is None or value == "":
        return value
    s = str(value).strip()
    if s and s[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + str(value)
    return value


@router.message(Command("admin"), IsAdminFilter())
async def cmd_admin(message: Message):
    """Админ-панель"""
    admin_text = """<b>🔐 Админ-панель</b>

Выберите действие:"""
    
    await message.answer(admin_text, reply_markup=get_admin_keyboard())


@router.callback_query(F.data == "admin_stats", IsAdminFilter())
async def admin_stats(callback: CallbackQuery):
    """Статистика по опросам"""
    async with aiosqlite.connect(get_db_path()) as db:
        # Общая статистика
        async with db.execute("SELECT COUNT(*) FROM surveys") as cursor:
            total_surveys = (await cursor.fetchone())[0]
        
        async with db.execute("SELECT COUNT(*) FROM users") as cursor:
            total_users = (await cursor.fetchone())[0]
        
        async with db.execute("SELECT COUNT(*) FROM referrals WHERE gift_claimed = 1") as cursor:
            gifts_claimed = (await cursor.fetchone())[0]
        
        # Пользователи, прошедшие тестирование
        async with db.execute("SELECT COUNT(DISTINCT user_id) FROM surveys") as cursor:
            users_completed = (await cursor.fetchone())[0]
        
        # Пользователи, которые были приглашены
        async with db.execute("SELECT COUNT(DISTINCT referred_id) FROM referrals") as cursor:
            users_referred = (await cursor.fetchone())[0]
        
        # Пользователи, которые пригласили других
        async with db.execute("SELECT COUNT(DISTINCT referrer_id) FROM referrals") as cursor:
            users_referrers = (await cursor.fetchone())[0]
    
    stats_text = f"""<b>📊 Статистика бота</b>

👥 Всего пользователей: {total_users}
✅ Прошли тестирование: {users_completed}
📝 Всего опросов: {total_surveys}
🎁 Подарков выдано: {gifts_claimed}
👥 Были приглашены: {users_referred}
👥 Пригласили других: {users_referrers}"""
    
    try:
        await callback.message.edit_text(stats_text, reply_markup=get_admin_keyboard())
    except TelegramBadRequest as e:
        if "message is not modified" not in (e.message or ""):
            raise
    await callback.answer()


@router.callback_query(F.data == "admin_users", IsAdminFilter())
async def admin_users(callback: CallbackQuery):
    """Список пользователей"""
    async with aiosqlite.connect(get_db_path()) as db:
        # Получаем всех пользователей с информацией о прохождении теста и приглашениях
        async with db.execute("""
            SELECT 
                u.user_id,
                u.full_name,
                u.username,
                CASE WHEN s.user_id IS NOT NULL THEN 1 ELSE 0 END as has_survey,
                CASE WHEN r1.referred_id IS NOT NULL THEN 1 ELSE 0 END as was_referred,
                CASE WHEN r2.referrer_id IS NOT NULL THEN 1 ELSE 0 END as has_referrals,
                r1.referrer_id,
                ref.full_name as referrer_name,
                ref.username as referrer_username
            FROM users u
            LEFT JOIN surveys s ON u.user_id = s.user_id
            LEFT JOIN referrals r1 ON u.user_id = r1.referred_id
            LEFT JOIN users ref ON r1.referrer_id = ref.user_id
            LEFT JOIN referrals r2 ON u.user_id = r2.referrer_id
            GROUP BY u.user_id
            ORDER BY u.created_at DESC
            LIMIT 50
        """) as cursor:
            users = await cursor.fetchall()
    
    if not users:
        await callback.message.edit_text(
            "<b>👥 Пользователи</b>\n\nПользователей пока нет.",
            reply_markup=get_admin_keyboard()
        )
        await callback.answer()
        return
    
    users_text = "<b>👥 Пользователи</b>\n\n"
    for user in users:
        user_id, full_name, username, has_survey, was_referred, has_referrals, referrer_id, referrer_name, referrer_username = user
        
        full_name = full_name or "Не указано"
        username = f"@{username}" if username else "Не указано"
        
        status_icons = []
        if has_survey:
            status_icons.append("✅ Тест")
        if was_referred:
            status_icons.append("👥 Приглашен")
        if has_referrals:
            status_icons.append("🎁 Пригласил")
        
        status = " ".join(status_icons) if status_icons else "❌ Нет активности"
        
        users_text += f"<b>{full_name}</b>\n"
        users_text += f"Username: {username}\n"
        users_text += f"ID: {user_id}\n"
        users_text += f"Статус: {status}\n"
        
        # Добавляем информацию о реферере, если пользователь был приглашен
        if was_referred and referrer_id:
            referrer_name = referrer_name or "Не указано"
            referrer_username = f"@{referrer_username}" if referrer_username else "Не указано"
            users_text += f"Пригласил: {referrer_name} ({referrer_username}, ID: {referrer_id})\n"
        
        users_text += "\n"
    
    if len(users) == 50:
        users_text += "\n<i>Показаны последние 50 пользователей</i>"
    
    await callback.message.edit_text(users_text, reply_markup=get_admin_keyboard())
    await callback.answer()


async def generate_excel_export() -> str:
    """Генерация Excel файла с результатами опросов (без персональных данных)"""
    logger.info("Начало генерации Excel файла")
    
    # Создаем временный файл
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()
    logger.info(f"Создан временный файл: {temp_file.name}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты опросов"
    
    # Заголовки
    headers = [
        "ID опроса",
        "Возраст",
        "Пол",
        "Рост (см)",
        "Вес (кг)",
        "ИМТ",
        "Образование",
        "Финансовая стабильность",
        "Курение",
        "Алкоголь (раз/неделю)",
        "Соль (г/день)",
        "Другие вредные привычки",
        "Время за экраном (ч/день)",
        "Физическая активность",
        "Ночные дежурства",
        "Ставка ночных дежурств",
        "Хронические заболевания",
        "Лекарства постоянно",
        "Семейный анамнез",
        "Уровень стресса (1-10)",
        "Качество сна (1-10)",
        "PHQ-2 балл",
        "PHQ-9 балл",
        "АД1 систолическое",
        "АД1 диастолическое",
        "АД1 время",
        "АД1 пульс",
        "АД2 систолическое",
        "АД2 диастолическое",
        "АД2 время",
        "АД2 пульс",
        "АД3 систолическое",
        "АД3 диастолическое",
        "АД3 время",
        "АД3 пульс",
        "Среднее АД систолическое",
        "Среднее АД диастолическое",
        "Модель тонометра",
        "Препараты (текст)",
        "Источник информации",
        "Балл риска",
        "Уровень риска",
        "Дата прохождения"
    ]
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Получаем данные из БД (с новыми полями: пульс, модель тонометра, препараты)
    async with aiosqlite.connect(get_db_path()) as db:
        try:
            async with db.execute("""
                SELECT 
                    id,
                    age, gender, height, weight, bmi,
                    education, financial_stability,
                    smoking, alcohol_per_week, salt_per_day, other_habits,
                    screen_time, physical_activity,
                    night_shifts, night_shifts_rate,
                    chronic_diseases, medications, family_history,
                    stress_level, sleep_quality,
                    phq2_score, phq9_score,
                    bp1_systolic, bp1_diastolic, bp1_time, bp1_pulse,
                    bp2_systolic, bp2_diastolic, bp2_time, bp2_pulse,
                    bp3_systolic, bp3_diastolic, bp3_time, bp3_pulse,
                    tonometer_model, medications_text,
                    referral_source, risk_score, risk_level,
                    completed_at
                FROM surveys
                ORDER BY completed_at DESC
            """) as cursor:
                rows = await cursor.fetchall()
            has_new_columns = True
        except aiosqlite.OperationalError:
            # Старая БД без новых колонок — выгружаем без них
            async with db.execute("""
                SELECT 
                    id,
                    age, gender, height, weight, bmi,
                    education, financial_stability,
                    smoking, alcohol_per_week, salt_per_day, other_habits,
                    screen_time, physical_activity,
                    night_shifts, night_shifts_rate,
                    chronic_diseases, medications, family_history,
                    stress_level, sleep_quality,
                    phq2_score, phq9_score,
                    bp1_systolic, bp1_diastolic, bp1_time,
                    bp2_systolic, bp2_diastolic, bp2_time,
                    bp3_systolic, bp3_diastolic, bp3_time,
                    referral_source, risk_score, risk_level,
                    completed_at
                FROM surveys
                ORDER BY completed_at DESC
            """) as cursor:
                rows = await cursor.fetchall()
            has_new_columns = False
    
    logger.info(f"Получено {len(rows)} записей из базы данных (новые колонки: {has_new_columns})")
    
    # Записываем данные
    if not rows:
        logger.warning("Нет данных для записи в Excel")
        wb.save(temp_file.name)
        return temp_file.name
    
    for row_idx, row in enumerate(rows, 2):
        if has_new_columns:
            # Индексы с новыми колонками: 23=bp1_sys ... 26=bp1_pulse, 27=bp2_sys ... 35=tonometer, 36=medications_text, 37=referral...
            bp1_sys, bp1_dia = row[23], row[24]
            bp2_sys, bp2_dia = row[27], row[28]
            bp3_sys, bp3_dia = row[31], row[32]
            avg_sys = (bp1_sys + bp2_sys + bp3_sys) // 3 if (bp1_sys and bp2_sys and bp3_sys) else None
            avg_dia = (bp1_dia + bp2_dia + bp3_dia) // 3 if (bp1_dia and bp2_dia and bp3_dia) else None
            data_row = [
                row[0], row[1] if row[1] else "", row[2] if row[2] else "", row[3] if row[3] else "", row[4] if row[4] else "",
                round(row[5], 2) if row[5] else "", row[6] if row[6] else "", row[7] if row[7] else "",
                "Да" if row[8] else "Нет", row[9] if row[9] is not None else 0, round(row[10], 1) if row[10] else "",
                "Да" if row[11] else "Нет", row[12] if row[12] else "", "Да" if row[13] else "Нет", "Да" if row[14] else "Нет",
                row[15] if row[15] else "", row[16] if row[16] else "", "Да" if row[17] else "Нет", "Да" if row[18] else "Нет",
                row[19] if row[19] else "", row[20] if row[20] else "", row[21] if row[21] else 0, row[22] if row[22] else 0,
                row[23] if row[23] else "", row[24] if row[24] else "", row[25] if row[25] else "", row[26] if row[26] else "",
                row[27] if row[27] else "", row[28] if row[28] else "", row[29] if row[29] else "", row[30] if row[30] else "",
                row[31] if row[31] else "", row[32] if row[32] else "", row[33] if row[33] else "", row[34] if row[34] else "",
                avg_sys if avg_sys else "", avg_dia if avg_dia else "",
                row[35] if row[35] else "", row[36] if row[36] else "",
                row[37] if row[37] else "", row[38] if row[38] else "", row[39] if row[39] else "", row[40] if row[40] else ""
            ]
        else:
            # Без новых колонок: bp1_time=25, bp2_sys=26, bp2_time=28, bp3_sys=29, bp3_time=31, referral=32, risk_score=33, risk_level=34, date=35
            bp1_sys, bp1_dia = row[23], row[24]
            bp2_sys, bp2_dia = row[26], row[27]
            bp3_sys, bp3_dia = row[29], row[30]
            avg_sys = (bp1_sys + bp2_sys + bp3_sys) // 3 if (bp1_sys and bp2_sys and bp3_sys) else None
            avg_dia = (bp1_dia + bp2_dia + bp3_dia) // 3 if (bp1_dia and bp2_dia and bp3_dia) else None
            data_row = [
                row[0], row[1] if row[1] else "", row[2] if row[2] else "", row[3] if row[3] else "", row[4] if row[4] else "",
                round(row[5], 2) if row[5] else "", row[6] if row[6] else "", row[7] if row[7] else "",
                "Да" if row[8] else "Нет", row[9] if row[9] is not None else 0, round(row[10], 1) if row[10] else "",
                "Да" if row[11] else "Нет", row[12] if row[12] else "", "Да" if row[13] else "Нет", "Да" if row[14] else "Нет",
                row[15] if row[15] else "", row[16] if row[16] else "", "Да" if row[17] else "Нет", "Да" if row[18] else "Нет",
                row[19] if row[19] else "", row[20] if row[20] else "", row[21] if row[21] else 0, row[22] if row[22] else 0,
                row[23] if row[23] else "", row[24] if row[24] else "", row[25] if row[25] else "",
                "",  # АД1 пульс
                row[26] if row[26] else "", row[27] if row[27] else "", row[28] if row[28] else "",
                "",  # АД2 пульс
                row[29] if row[29] else "", row[30] if row[30] else "", row[31] if row[31] else "",
                "",  # АД3 пульс
                avg_sys if avg_sys else "", avg_dia if avg_dia else "",
                "", "",  # Модель тонометра, Препараты (текст)
                row[32] if row[32] else "", row[33] if row[33] else "", row[34] if row[34] else "", row[35] if row[35] else ""
            ]
        
        for col, value in enumerate(data_row, 1):
            ws.cell(row=row_idx, column=col, value=_excel_safe_value(value))
    
    # Настраиваем ширину колонок
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # Сохраняем файл
    try:
        wb.save(temp_file.name)
        logger.info(f"Excel файл успешно сохранен: {temp_file.name}")
    except Exception as e:
        logger.error(f"Ошибка при сохранении Excel файла: {e}", exc_info=True)
        raise
    
    return temp_file.name


@router.callback_query(F.data == "admin_export", IsAdminFilter())
async def admin_export(callback: CallbackQuery):
    """Выгрузка результатов опросов в Excel"""
    await callback.answer("⏳ Проверяю данные...")
    
    # Сначала проверяем, есть ли данные
    async with aiosqlite.connect(get_db_path()) as db:
        async with db.execute("SELECT COUNT(*) FROM surveys") as cursor:
            count_row = await cursor.fetchone()
            survey_count = count_row[0] if count_row else 0
    
    if survey_count == 0:
        await callback.message.answer("ℹ️ <b>Пока никто не проходил тест</b>\n\nРезультаты появятся здесь после того, как пользователи пройдут опрос.")
        await callback.answer()
        return
    
    await callback.answer("⏳ Генерирую файл...")
    
    try:
        logger.info("Начало генерации Excel файла для админа")
        # Генерируем Excel файл
        file_path = await generate_excel_export()
        
        if not os.path.exists(file_path):
            logger.error(f"Файл не был создан: {file_path}")
            await callback.answer("❌ Ошибка: файл не был создан", show_alert=True)
            return
        
        logger.info(f"Отправка файла админу: {file_path}")
        # Отправляем файл
        file = FSInputFile(file_path, filename=f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        await callback.message.answer_document(
            document=file,
            caption="📥 <b>Результаты опросов</b>\n\nФайл содержит все ответы и результаты без персональных данных."
        )
        
        logger.info("Файл успешно отправлен админу")
        
        # Удаляем временный файл
        try:
            os.unlink(file_path)
            logger.info(f"Временный файл удален: {file_path}")
        except Exception as e:
            logger.warning(f"Не удалось удалить временный файл {file_path}: {e}")
        
        await callback.answer("✅ Файл успешно сгенерирован")
    except Exception as e:
            await callback.answer("❌ Ошибка при генерации файла", show_alert=True)
            logger.error(f"Ошибка при генерации Excel файла: {e}", exc_info=True)


@router.message(Command("send_unclaimed_gifts"), IsAdminFilter())
async def cmd_send_unclaimed_gifts(message: Message):
    """Отправить подарки всем реферерам, чьи приглашённые прошли тест, но подарок ещё не был выдан."""
    from app.handlers.survey_router import send_gift_to_referrer
    status = await message.answer("⏳ Ищу невыданные подарки...")
    sent = 0
    errors = 0
    async with aiosqlite.connect(get_db_path()) as db:
        async with db.execute("""
            SELECT r.referrer_id, r.referred_id
            FROM referrals r
            INNER JOIN surveys s ON s.user_id = r.referred_id
            WHERE r.gift_claimed = 0
            ORDER BY r.referrer_id, r.referred_id
        """) as cursor:
            rows = await cursor.fetchall()
    for referrer_id, referred_id in rows:
        try:
            await send_gift_to_referrer(message.bot, referred_id)
            sent += 1
            await asyncio.sleep(0.5)
        except Exception as e:
            logger.exception(f"Ошибка отправки подарка рефереру {referrer_id} за referred_id {referred_id}")
            errors += 1
    try:
        await status.edit_text(
            f"✅ <b>Готово</b>\n\n"
            f"Обработано записей: {len(rows)}\n"
            f"Подарков отправлено: {sent}\n"
            f"Ошибок: {errors}"
        )
    except Exception:
        await message.answer(
            f"✅ Обработано: {len(rows)}, отправлено: {sent}, ошибок: {errors}"
        )


@router.message(Command("test_gift"), IsAdminFilter())
async def cmd_test_gift(message: Message):
    """Команда для тестирования выбора подарка и генерации реферальной ссылки"""
    user_id = message.from_user.id
    
    # Проверяем, существует ли пользователь в БД
    async with aiosqlite.connect(get_db_path()) as db:
        async with db.execute("SELECT user_id FROM users WHERE user_id = ?", (user_id,)) as cursor:
            user_exists = await cursor.fetchone()
        
        if not user_exists:
            # Создаем пользователя, если его нет
            from app.utils.referral import generate_referral_code
            ref_code = generate_referral_code(user_id)
            await db.execute(
                "INSERT INTO users (user_id, username, full_name, referral_code) VALUES (?, ?, ?, ?)",
                (user_id, message.from_user.username or "Admin", message.from_user.full_name or "Admin", ref_code)
            )
            await db.commit()
            logger.info(f"Создан пользователь {user_id} для тестирования")
    
    gift_text = """🎁 <b>Тестирование реферальной системы</b>

Выберите подарок для генерации реферальной ссылки:"""
    
    sent_message = await message.answer(gift_text, reply_markup=get_gift_keyboard())
    
    # Сохраняем ID сообщения для проверки в обработчике
    # Можно использовать состояние или просто проверить, что это админ


@router.callback_query(F.data.startswith("gift_"), IsAdminFilter())
async def admin_process_gift_choice(callback: CallbackQuery, state: FSMContext):
    """Обработчик выбора подарка для админа (тестирование)"""
    # Проверяем, что пользователь НЕ в состоянии waiting_gift_choice (чтобы не конфликтовать с основным обработчиком)
    current_state = await state.get_state()
    if current_state == SurveyStates.waiting_gift_choice:
        # Если админ в состоянии waiting_gift_choice, пропускаем этот обработчик
        # Основной обработчик в survey_router обработает это
        return
    
    gift_map = {
        "gift_nutrition": "Питание",
        "gift_activity": "Физическая активность",
        "gift_stress": "Стресс"
    }
    gift_type = gift_map.get(callback.data)
    
    if not gift_type:
        await callback.answer("❌ Неизвестный подарок", show_alert=True)
        return
    
    user_id = callback.from_user.id
    
    # Сохраняем предпочтение подарка
    async with aiosqlite.connect(get_db_path()) as db:
        await db.execute(
            "UPDATE users SET preferred_gift = ? WHERE user_id = ?",
            (gift_type, user_id)
        )
        await db.commit()
    
    # Генерируем реферальную ссылку
    bot_username = (await callback.message.bot.get_me()).username
    ref_link = f"https://t.me/{bot_username}?start={user_id}"
    
    result_text = f"""✅ <b>Подарок выбран!</b>

Вы выбрали подарок: <b>{gift_type}</b>

🔗 <b>Ваша реферальная ссылка:</b>
<code>{ref_link}</code>

📋 <b>Для тестирования:</b>
1. Скопируйте ссылку
2. Откройте её в другом аккаунте (или попросите друга)
3. Пройдите тест
4. После завершения теста вы получите подарок!

Или используйте команду /test_complete для симуляции прохождения теста."""
    
    await callback.message.edit_text(result_text)
    await callback.answer(f"✅ Выбрано: {gift_type}")


@router.message(Command("test_complete"), IsAdminFilter())
async def cmd_test_complete(message: Message):
    """Команда для симуляции прохождения теста (создает запись в surveys)"""
    user_id = message.from_user.id
    
    # Проверяем, существует ли пользователь в БД
    async with aiosqlite.connect(get_db_path()) as db:
        async with db.execute("SELECT user_id FROM users WHERE user_id = ?", (user_id,)) as cursor:
            user_exists = await cursor.fetchone()
        
        if not user_exists:
            await message.answer("❌ Пользователь не найден в БД. Сначала используйте /test_gift для создания пользователя.")
            return
        
        # Проверяем, есть ли уже запись о прохождении теста
        async with db.execute("SELECT id FROM surveys WHERE user_id = ?", (user_id,)) as cursor:
            survey_exists = await cursor.fetchone()
        
        if survey_exists:
            # Обновляем существующую запись
            await db.execute("""
                UPDATE surveys 
                SET completed_at = CURRENT_TIMESTAMP
                WHERE user_id = ?
            """, (user_id,))
            await db.commit()
            await message.answer("✅ Запись о прохождении теста обновлена (время изменено на текущее)")
        else:
            # Создаем новую запись с минимальными данными для тестирования
            await db.execute("""
                INSERT INTO surveys (
                    user_id, consent, age, gender, height, weight, bmi, 
                    education, financial_stability, smoking, alcohol_per_week, 
                    salt_per_day, other_habits, screen_time, physical_activity,
                    night_shifts, night_shifts_rate, chronic_diseases, medications, medications_text,
                    family_history, stress_level, sleep_quality, phq2_score, phq9_score,
                    bp1_systolic, bp1_diastolic, bp1_pulse, bp1_time,
                    bp2_systolic, bp2_diastolic, bp2_pulse, bp2_time,
                    bp3_systolic, bp3_diastolic, bp3_pulse, bp3_time,
                    tonometer_model, referral_source, risk_level, risk_score
                ) VALUES (?, 1, 30, 'мужской', 175, 75, 24.5, 
                    'Высшее', 'Средняя', 0, 0, 5.0, 0, 4, 1,
                    0, NULL, 'нет', 0, NULL, 0, 5, 7, 0, 0,
                    120, 80, 72, '12:00',
                    125, 82, 75, '16:00',
                    118, 78, 70, '20:00',
                    NULL, 'Другое', 'низкий', 5)
            """, (user_id,))
            await db.commit()
            await message.answer("✅ Создана запись о прохождении теста (тестовые данные)")
        
        # Проверяем, был ли пользователь приглашен, и отправляем подарок рефереру
        from app.handlers.survey_router import send_gift_to_referrer
        await send_gift_to_referrer(message.bot, user_id)
        
        await message.answer("✅ Тест пройден! Проверьте, был ли отправлен подарок рефереру (если он был указан).")
